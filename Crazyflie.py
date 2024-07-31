# Crazyflie Import
import cflib.crtp
from cflib.utils import uri_helper
from cflib.crazyflie import Crazyflie
from cflib.crazyflie.syncCrazyflie import SyncCrazyflie
from cflib.positioning.position_hl_commander import PositionHlCommander
from cflib.crazyflie.log import LogConfig

# Python Import
import time
import numpy as np
from openpyxl import Workbook
from openpyxl import load_workbook

class CrazyFlie:
    """
    CrazyFlie: Creates, flies, and tracks the crazyflie (cf). When flight is complete, creates a log to  
    store transit data.
    On input:
        uriName (int): radio number
        cam (object): camera object for third person tracking
    On output:
        N/A
        *CFLocData (.xlsx file): file with transit data, both camera and telemitry data.
    Call:
       Crazyflie.CrazyFlie(uri , rsCam)
    Author:   
        W.Raley    
        UU
        Summer 2024
    """
    def __init__ (self , uriNum , cam, inst):
        # Storage for location data
        self.telReportedLoc = []
        self.camReportedLoc = []
        self.transitData = []
        self.transitionPts = []
        self.route = inst[: , 0:3]

        self.position_estimate = []

        # radio information
        self.cfNum = uriNum
        uriName = 'radio://0/80/2M/E7E7E7E70' + self.cfNum
        self.uri = uri_helper.uri_from_env(default= uriName)

        # Camera information
        self.cam = cam

        # Current state of cf
        self.state = True

        # Speed of flight
        self.speed = inst[0,7]

    def createLog(self):
            # prepares log data
            logconf = LogConfig(name = 'Position', period_in_ms=10)
        
            logconf.add_variable('kalman.stateX' , 'float')
            logconf.add_variable('kalman.stateY' , 'float')
            logconf.add_variable('kalman.stateZ' , 'float')
       
            self.scf.cf.log.add_config(logconf)
            # prints position data
            logconf.data_received_cb.add_callback(self.log_pos_callback)

            return logconf

    # Sets the current position
    def setPos(self , logConf, pc):
        # Starts the log
        logConf.start()
        time.sleep(0.1)

       # self.initX = pc._x
        #self.initY = pc._y
       # self.initZ = pc._z

        logConf.stop()
        logConf.delete()
        
        return self.createLog()
    
    def cfState(self):

        return self.state

    def updateLoc(self):
        
        # Updates telemetry data
        locArr = self.position_estimate
        self.telReportedLoc = np.append(self.telReportedLoc , locArr )
        
        # Updates Camera reported location
        camReported = self.cam.locateCF(self.cfNum , locArr)
    
        if camReported[0] == 0 or camReported[1] == 0 or camReported[2] == 0 or abs(camReported[0]) > 3:
            while camReported[0] == 0 or camReported[1] == 0 or camReported[2] == 0 or abs(camReported[0]) > 5:
                 camReported = self.cam.locateCF(self.cfNum , locArr)

        self.camReportedLoc = np.append(self.camReportedLoc , [camReported[0] , camReported[2] ,camReported[1]])
        

    def fly(self):

        with PositionHlCommander(self.scf) as pc:
            
            # Sets Position
            logConf = self.createLog()
            self.logConf = self.setPos(logConf , pc) 
            
            initX = pc._x
            initY = pc._y
            initZ = pc._z

            self.transitData = np.append(self.transitData , [initX, initY, initZ])
            
            # Sets flying status to off
            pc._is_flying = False

            self.logConf.start()

            # move
            pc.take_off(height=0.1 , velocity=self.speed)
            self.transitionPts = np.append(self.transitionPts , len(self.camReportedLoc))

            [currX , currY , currZ ] = self.directTransit(self.route[0 , :], [initX, initY, initZ], pc)
            np.delete(self.route , 0)

            for pt in self.route:
                
               [currX , currY , currZ ] = self.directTransit(pt, [currX , currY , currZ ], pc)

            # landing sequence
            pc._is_flying = True
            pc.land(velocity = self.speed)
            self.logConf.stop()
            self.transitData = np.append(self.transitData , [pc._x , pc._y  , pc._z ])
            self.state = False

            self.recordData()

    def directTransit(self , goToArr, currLoc, pc):

        #xUpdate = currLoc[0] + goToArr[0]
        #yUpdate = currLoc[1] + goToArr[1]
        #zUpdate = currLoc[2] + goToArr[2]

        self.transitData = np.append(self.transitData , [currLoc[0] , currLoc[1] , currLoc[2]])
       
        pc.go_to(goToArr[0], goToArr[1], goToArr[2], velocity = self.speed)
        time.sleep(1)
        self.transitionPts = np.append(self.transitionPts , len(self.camReportedLoc))

        return [goToArr[0], goToArr[1], goToArr[2]]



    # Sets location to current estimate
    def log_pos_callback(self , timestamp, data, logconf):
        
        # print(data)
        
        tempEstimate = [0, 0, 0]

        tempEstimate[0] = data['kalman.stateX']
        tempEstimate[1] = data['kalman.stateY']
        tempEstimate[2] = data['kalman.stateZ']

        self.position_estimate = tempEstimate

    # Records data to spreadsheet
    def recordData(self):
        
        filename = 'CFLocData.xlsx'
        self.createWB(filename)
        doc = load_workbook(filename)
        
        sheet = doc.active
    
        for i in range(int(len(self.telReportedLoc) / 3)) :
            ii = i * 3
            line = [self.telReportedLoc[ii] , self.telReportedLoc[ii - 2] , self.telReportedLoc[ii - 1]]
            sheet.append(line)

        doc.save(filename)
        
        sheet = doc.create_sheet("Cam Data")

        for i in range(int(len(self.camReportedLoc) / 3)) :
            ii = i * 3
            line = [self.camReportedLoc[ii] , self.camReportedLoc[ii - 1] , self.camReportedLoc[ii - 2]]
            sheet.append(line)
        

        doc.save(filename)

        sheet = doc.create_sheet("Transit Data")

        for ii in range(int(len(self.route))) :

            line = [self.route[ii , 0] , self.route[ii , 1] , self.route[ii , 2]]
            sheet.append(line)

        doc.save(filename)

        sheet = doc.create_sheet("Transition Points")

        for i in range(len(self.transitionPts)):
            line = [self.transitionPts[i]]
            sheet.append(line)

        doc.save(filename)
    
    # Creates spreadsheet to record data
    def createWB(self, fileName):
        wb = Workbook()
        wb.create_sheet("Telemetry Data")
        del wb['Sheet']
        wb.save(fileName)
        self.wb = wb

    def run(self):

        cflib.crtp.init_drivers()

        with SyncCrazyflie(self.uri , cf=Crazyflie(rw_cache='./cache')) as scf:
           
            # save scf for access throughout class
            self.scf = scf
    
            self.fly()

        